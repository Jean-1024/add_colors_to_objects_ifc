# After running this script, save it to a new file then open the new file you just saved to check the color result

from openpyxl import load_workbook
import ifcopenshell
import ifcopenshell.util.element
from bonsai.bim.ifc import IfcStore
import bonsai.tool as tool
import re
import bpy

# --------------------- CONFIG ---------------------
XLSX_PATH = r"SimpleBIM_Type_Filter.xlsx"
SHEET_NAME = "ModelView"
# --------------------------------------------------

def argb_to_rgb(argb_hex_string):
    # '#AARRGGBB' -> (R,G,B) in 0~1
    # If the first char in the string is #, remove it
    if argb_hex_string.startswith("#"):
        argb_hex_string = argb_hex_string[1:]
    if len(argb_hex_string) != 8:
        raise ValueError("Invalid argb hex string format. Expected #AARRGGBB.")
    r = int(argb_hex_string[2:4], 16) / 255.0
    g = int(argb_hex_string[4:6], 16) / 255.0
    b = int(argb_hex_string[6:8], 16) / 255.0
    return (r, g, b)

def get_excel_mapping(xlsx_path, sheet_name):
    # Returns { system_type_lower: (r,g,b) }
    workbook = load_workbook(xlsx_path)
    worksheet = workbook[sheet_name]
    start_row = None
    name_column = None
    color_column = None

    for row in worksheet.iter_rows():
        for cell in row:
            # Find the cell that has string 'Color'
            if cell.value == "Color":
                start_row = cell.row + 1
                name_column = cell.column - 1
                color_column = cell.column
                break
        if start_row is not None:
            break

    mapping = {}
    if start_row is None or name_column is None or color_column is None:
        print("'Color' header not found.")
        return mapping

    for i in range(start_row, worksheet.max_row + 1):
        name_cell = worksheet.cell(row=i, column=name_column)
        color_cell = worksheet.cell(row=i, column=color_column)
        # .strip() — removes leading and trailing whitespace characters (spaces, tabs, newlines)
        name = (name_cell.value or "").strip()
        if not name:
            continue
        
        # when a cell has no fill, Excel often represents the color as either:
        # - None
        # - "00000000" (transparent, a=00 is fully transparent, rgb=000000 is black, we see a white cell bc background is white)
        # - or sometimes "FFFFFFFF" (white) depending on theme/formatting.
        # (color black is #FF000000)
        # if the cell has no value or an empty value, continue.
        fill = color_cell.fill
        if not fill or not fill.start_color:
            continue
        argb_hex = fill.start_color.rgb
        if not argb_hex or argb_hex in ("00000000"):
            continue
        r, g, b = argb_to_rgb(argb_hex)
        # use .lower() to ensures case-insensitive matching when looking up colors later
        mapping[str(name).lower()] = (r, g, b)

    return mapping

def get_body_item(prod):
    # Purpose: (Assume single-item body representation) Return the representation item for the product's Body representation.
    
    # get the attribute named "Representation" from the object prod.
    # prod.Representation → is usually an IfcProductDefinitionShape
    # prod.Representation.Representations → is a list of IfcShapeRepresentation objects
    # If prod has an attribute Representation, its value is stored in rep_count.
    # If prod does not have that attribute, getattr() returns the default value None.
    # rep_count is either: an IfcProductDefinitionShape instance, or None.
    rep_count = getattr(prod, "Representation", None)
    # Check whether Representation exists at all, and 
    # whether that Representation has a sub-attribute called Representations (which should be a list of shape representations)
    # If either prod.Representation or prod.Representation.Representations doesn’t exist or is empty/None, it returns an empty list.
    if not rep_count or not getattr(rep_count, "Representations", None):
        return []    
    
    # Get the list of shape representations, or an empty list if none exist
    representations = rep_count.Representations or []

    # Try to find the "Body" representation — the main 3D geometry (Assume only 1 Body representation for an object)
    body_representation = None

    # Loop through the list of 'IfcShapeRepresentation's 
    for rep in representations:
        if getattr(rep, "RepresentationIdentifier", None) == "Body":
            body_representation = rep
            break  # Stop once we find the "Body" representation

    # If nothing found, return an empty list (no usable geometry)
    if not body_representation:
        return []

    items = getattr(body_representation, "Items", None) or []
    # Return the first (and assumed only) item, or None if empty
    return items[0] if items else None

def get_or_make_psa(ifc, cache, r, g, b, transparency=0.0):
    # Cache by RGB(+transparency) to reuse styles across objects
    # cache format: {(r1, g1, b1, t1): psa1, (r2, g2, b2, t2): psa2}
    key = (round(r, 6), round(g, 6), round(b, 6), round(transparency, 6))
    # key = (r,g,b,t)
    if key in cache:
        # return psa, which is #6741=IfcPresentationStyleAssignment((#6740))
        return cache[key]

    rgb   = ifc.create_entity("IfcColourRgb", None, r, g, b)
    render = ifc.create_entity(
        "IfcSurfaceStyleRendering",
        rgb,                                    # SurfaceColour
        transparency,                           # Transparency (0 = opaque)
        None, None, None, None, None, None,     # (optional reflectance params)
        "NOTDEFINED"                            # ReflectanceMethod
    )
    surf  = ifc.create_entity("IfcSurfaceStyle", None, "BOTH", [render])
    psa   = ifc.create_entity("IfcPresentationStyleAssignment", [surf])
    
    # update cache
    cache[key] = psa
    return psa


def assign_style_to_item_instance(ifc, item, style):
    # Assign style to a specific representation item (instance-level) by creating/updating an IfcStyledItem on that item.
    
    '''
    In IFC, styling information (like color, materials, etc.) is often stored in an IfcStyledItem entity.
    For any geometric item (e.g., IfcFacetedBrep, IfcExtrudedAreaSolid, IfcShell, etc.), there’s an inverse attribute called StyledByItem that lists the styles applied to it.
    That’s a list-like inverse relationship — meaning an item can have zero or more style definitions associated with it.
    Every representation item can have 0, 1, or several styles.
    Each IfcStyledItem can hold one or more IfcPresentationStyle objects (like surface colors, curves, or text styles).
    In most real IFC files:
        Each geometry item (like a solid, face, or curve) has at most one IfcStyledItem.
        That single IfcStyledItem will include all relevant style information (in its .Styles list).
    '''
    
    existing = getattr(item, "StyledByItem", None)
    # If there’s already one IfcStyledItem attached, it reuses it rather than creating a duplicate.
    # Each geometry item only has one styled item
    '''
    existing: Makes sure the attribute actually exists, some IFC classes don’t have StyledByItem, or it could be None
    len(existing) > 0: Makes sure there’s at least one style in the list, many items don’t have any styling yet, so the list could be empty
    existing[0]: Ensures the first element is a valid IfcStyledItem object, sometimes IFC files have a placeholder or a None entry in the list
    
    styled: This is an IfcStyledItem instance.
    Represents a link between a geometric item (e.g., a face, solid, or curve) and one or more style definitions.
    'styled.Styles': This is a list-like attribute of IfcStyledItem.
    Each element is an IfcPresentationStyleAssignment (or a subtype, e.g., IfcSurfaceStyle, IfcCurveStyle).
    These define how the item should look (color, transparency, line weight, etc.).
    '''
    
    if existing and len(existing) > 0 and existing[0]:
        styled = existing[0]
        # It reads the existing list of IfcPresentationStyle references from styled.Styles
        # If the item already has a valid IfcStyledItem, update it by adding the new style.
        styles = list(styled.Styles or [])
        '''
        styles  --> [ IfcSurfaceStyle1, IfcSurfaceStyle2, ... ]
        IfcSurfaceStyle
             ├─ Name: "RedPaint"
             ├─ Side: "BOTH"
             └─ Styles: [IfcSurfaceStyleRendering]
                   ├─ SurfaceColour: RGB(1.0, 0.0, 0.0)
                   ├─ Transparency: 0.0
                   └─ ReflectanceMethod: "BLINN"
        styles[0] is an IfcSurfaceStyle, 
        styles[0].Styles[0] is an IfcSurfaceStyleRendering
        '''

        if style not in styles:
            styles.append(style)
            styled.Styles = styles
        return styled
    else:
        # Create new IfcStyledItem bound to the item
        return ifc.create_entity("IfcStyledItem", Item=item, Styles=[style], Name=None)

def main():
    ifc = IfcStore.get_file()
    if not ifc:
        raise RuntimeError("No IFC open.")

    mapping = get_excel_mapping(XLSX_PATH, SHEET_NAME)
    #print("mapping: ", mapping)
    if not mapping:
        print("[WARN] No mapping. End the program.")
        return
    
    # Set up a list 'style_cache = {}' to store the unique color mapping we can find when we loop through all the objects
    # Each time we get a new 'name--color' that does not exist in cache, we add it to cache to avoid getting duplicate colors
    # format: {(r1, g1, b1, t1): psa1, (r2, g2, b2, t2): psa2}
    style_cache = {}
    # counting how many items we assign color for
    assigned_items = 0
    # counting how many objects we assign color for
    assigned_objects = 0
    # counting how many object that does not have body representation
    no_body = 0
    # counting the color names that have no corresponding color are found in the mapping we get from excel sheet
    no_color = set()

        # Loop iterates over Blender-side objects, not IFC-side objects
        # Each Blender object has metadata linking it back to its IFC entity
    for obj in bpy.data.objects:
        #print("\n object: ", obj)
        #object:  <bpy_struct, Object("IfcAirTerminal/Return Diffuser:600 x 600 Face 300 x 300 Conne") at 0x32612ac08>

        # Find corresponding IFC entity from each object
        entity = tool.Ifc.get_entity(obj)
        #print("\n entity: ", entity)
        #entity:  #359=IfcAirTerminal('0G1L222Lj83B_4a2DpVHkP',#18,'Return Diffuser:600 x 600 Face 300 x 300 Connection:1350981',$,'Return Diffuser:600 x 600 Face 300 x 300 Connection',#358,#355,'1350981',.NOTDEFINED.)
        
        if not entity:
            continue

        # Only work on instances that have 7-digits in Name
        if not re.search(r'\b\d{7}\b', (entity.Name or "")):
            continue
         
        # Get all the psets each object has
        psets = ifcopenshell.util.element.get_psets(entity, psets_only=True) or {}
        # Get the values (color name) on 'System Type' from the pset 'Mechanical' for each object
        system_type = (psets.get("Mechanical", {}) or {}).get("System Type")
        #print("\nname:", entity.Name)
        #print("system type value:", system_type)
        if not system_type:
            continue

        key = (str(system_type) or "").strip().lower()
        rgb = mapping.get(key)
        #print("lookup key:", key, "rgb:", rgb)
        # If no rgb color are found in mapping for the key (color name), save the key value to 'no_color' and later print out for checking.
        if not rgb:
            no_color.add(system_type)
            continue

        r, g, b = rgb
        style = get_or_make_psa(ifc, style_cache, r, g, b, transparency=0.0)
        #print("style_cache:", style_cache)
        #style_cache: {(1.0, 0.0, 1.0, 0.0): #6749=IfcPresentationStyleAssignment((#6748)), (1.0, 0.0, 0.0, 0.0): #6753=IfcPresentationStyleAssignment((#6752))}
        #print("style:", style)
        #style: #6741=IfcPresentationStyleAssignment((#6740))
        #style: #6745=IfcPresentationStyleAssignment((#6744))
        
        # Assume single-item body representation
        item = get_body_item(entity)
        if not item:
            no_body += 1
            continue
        if item:
            assign_style_to_item_instance(ifc, item, style)
            assigned_items += 1
        assigned_objects += 1


    print(f"[DONE] Styled {assigned_items} representation items.")
    print(f"[DONE] Styled {assigned_objects} representation objects.")
    if no_body:
        print(f"[INFO] {no_body} products had no Body representation or items.")
    if no_color:
        print(f"[INFO] Missing Excel colors for System Types: {sorted(set(no_color))}")

if __name__ == "__main__":
    main()
